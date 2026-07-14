from concurrent.futures import ThreadPoolExecutor, as_completed
from os import listdir, path

from bs4 import BeautifulSoup, Tag
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from constants.constants import ENCODING, EXPORT_FOLDER, SAVE_FOLDER
from domain.exceptions import FileSystemOperationError, ScheduleConversionError


class ExcelFetcher:
    """
    Converts downloaded Plansoft HTML table files into Excel spreadsheets.
    """

    @staticmethod
    def _parse_html_table(table_tag: Tag) -> list[list[str]]:
        rows = table_tag.find_all("tr")

        grid: list[list[str]] = []
        rowspan_tracker: dict[tuple[int, int], str] = {}

        for row_idx, row in enumerate(rows):
            cells = row.find_all(["td", "th"])
            col_idx = 0
            grid_row: list[str] = []

            while col_idx < 100:
                key = (row_idx, col_idx)

                if key in rowspan_tracker:
                    grid_row.append(rowspan_tracker.pop(key))
                    col_idx += 1
                    continue

                if not cells:
                    break

                cell = cells.pop(0)
                value = cell.get_text(strip=True)

                col_span = int(cell.get("colspan", 1))
                row_span = int(cell.get("rowspan", 1))

                for _ in range(col_span):
                    grid_row.append(value)

                    if row_span > 1:
                        for row_offset in range(1, row_span):
                            rowspan_tracker[(row_idx + row_offset, col_idx)] = value

                    col_idx += 1

            grid.append(grid_row)

        max_width = max((len(row) for row in grid), default=0)

        for row in grid:
            row.extend([""] * (max_width - len(row)))

        return grid

    @staticmethod
    def _adjust_column_widths(ws: Worksheet) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    @staticmethod
    def save_to_excel(data: list[list[str]], filename: str) -> None:
        try:
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Schedules"

            for row_idx, row in enumerate(data, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            ExcelFetcher._adjust_column_widths(ws)
            workbook.save(filename)
        except Exception as exc:
            raise ScheduleConversionError(
                "Cannot save converted Plansoft Excel file.",
                details={
                    "filename": filename,
                    "reason": str(exc),
                },
            ) from exc

    def _process_single_html_file(self, filepath: str, output_path: str) -> bool:
        try:
            with open(filepath, "r", encoding=ENCODING) as file:
                soup = BeautifulSoup(file, "html.parser")
                table = soup.find("table")
        except OSError as exc:
            raise FileSystemOperationError(
                "Cannot read Plansoft HTML file.",
                details={
                    "filepath": filepath,
                    "reason": str(exc),
                },
            ) from exc

        if not isinstance(table, Tag):
            print(f"No table in: {path.basename(filepath)}")
            return False

        data = self._parse_html_table(table)
        self.save_to_excel(data, output_path)
        print(f"Saved: {output_path}")
        return True

    def html_to_excel_parallel(self) -> None:
        try:
            files = [filename for filename in listdir(SAVE_FOLDER) if filename.endswith(".htm")]
        except OSError as exc:
            raise FileSystemOperationError(
                "Cannot list Plansoft HTML files.",
                details={
                    "save_folder": SAVE_FOLDER,
                    "reason": str(exc),
                },
            ) from exc

        print(f"Converting {len(files)} HTML files to Excel...")

        failures = 0

        with ThreadPoolExecutor(max_workers=16) as executor:
            futures = []

            for filename in files:
                filepath = path.join(SAVE_FOLDER, filename)
                group_name = filename.rsplit(".", 1)[0]
                output_path = path.join(EXPORT_FOLDER, f"{group_name}.xlsx")
                futures.append(executor.submit(self._process_single_html_file, filepath, output_path))

            for future in as_completed(futures):
                try:
                    converted = future.result()
                except (FileSystemOperationError, ScheduleConversionError):
                    raise
                except Exception as exc:
                    raise ScheduleConversionError(
                        "Unexpected Plansoft HTML conversion failure.",
                        details={"reason": str(exc)},
                    ) from exc

                if not converted:
                    failures += 1

        if files and failures == len(files):
            raise ScheduleConversionError(
                "No Plansoft HTML files contained parseable tables.",
                details={"files_checked": len(files)},
            )
